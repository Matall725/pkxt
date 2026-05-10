import csv
import json
from collections import defaultdict
from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.db.models import Max, Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.db.models.functions import TruncDate

from notes.models import SessionNote
from payments.models import PaymentEntry, Receivable
from reminders.models import ReminderTask
from schedules.models import Schedule
from students.models import Student


ZERO_DECIMAL = Decimal("0")
WEEKDAY_LABELS = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]


def _build_finance_chart(today):
    month_start = today.replace(day=1)
    due_rows = (
        Receivable.objects.filter(issue_date__gte=month_start, issue_date__lte=today)
        .values("issue_date")
        .annotate(total=Sum("amount_due"))
        .order_by("issue_date")
    )
    received_rows = (
        PaymentEntry.objects.filter(received_at__date__gte=month_start, received_at__date__lte=today)
        .annotate(day=TruncDate("received_at"))
        .values("day")
        .annotate(total=Sum("amount"))
        .order_by("day")
    )

    due_map = defaultdict(lambda: ZERO_DECIMAL)
    received_map = defaultdict(lambda: ZERO_DECIMAL)
    for row in due_rows:
        due_map[row["issue_date"]] = row["total"] or ZERO_DECIMAL
    for row in received_rows:
        received_map[row["day"]] = row["total"] or ZERO_DECIMAL

    labels = []
    due_series = []
    received_series = []
    outstanding_series = []
    cursor = month_start
    cumulative_due = ZERO_DECIMAL
    cumulative_received = ZERO_DECIMAL

    while cursor <= today:
        cumulative_due += due_map[cursor]
        cumulative_received += received_map[cursor]
        outstanding = cumulative_due - cumulative_received
        if outstanding < ZERO_DECIMAL:
            outstanding = ZERO_DECIMAL

        labels.append(f"{cursor.month}/{cursor.day}")
        due_series.append(float(cumulative_due))
        received_series.append(float(cumulative_received))
        outstanding_series.append(float(outstanding))
        cursor += timedelta(days=1)

    return {
        "labels": json.dumps(labels, ensure_ascii=False),
        "due": json.dumps(due_series),
        "received": json.dumps(received_series),
        "outstanding": json.dumps(outstanding_series),
    }


def _build_student_overview():
    queryset = (
        Student.objects.prefetch_related("service_plans")
        .annotate(last_class_at=Max("schedules__start_at"))
        .order_by("-last_class_at", "-updated_at", "name")[:6]
    )

    rows = []
    for student in queryset:
        plan = student.active_service_plan
        rows.append(
            {
                "student": student,
                "plan": plan,
                "remaining_hours": getattr(plan, "remaining_hours", None),
                "owed_hours": getattr(plan, "owed_hours", None),
                "last_class_at": student.last_class_at,
            }
        )
    return rows


@login_required
def home(request):
    today = timezone.localdate()
    month_start = today.replace(day=1)
    yesterday = today - timedelta(days=1)

    today_schedules = list(
        Schedule.objects.select_related("student", "service_plan")
        .filter(start_at__date=today)
        .order_by("start_at")
    )
    yesterday_schedule_count = Schedule.objects.filter(start_at__date=yesterday).count()

    open_receivables = list(
        Receivable.objects.select_related("student", "service_plan")
        .filter(status__in=[Receivable.Status.PENDING, Receivable.Status.PARTIAL])
        .order_by("due_date", "id")
    )

    open_due_total = sum((receivable.amount_due for receivable in open_receivables), ZERO_DECIMAL)
    open_received_total = sum((receivable.amount_received for receivable in open_receivables), ZERO_DECIMAL)
    outstanding_total = open_due_total - open_received_total
    if outstanding_total < ZERO_DECIMAL:
        outstanding_total = ZERO_DECIMAL

    active_students_count = Student.objects.filter(status=Student.Status.ACTIVE).count()
    lead_students_count = Student.objects.filter(status=Student.Status.LEAD).count()

    month_received_total = (
        PaymentEntry.objects.filter(received_at__date__gte=month_start, received_at__date__lte=today)
        .aggregate(total=Sum("amount"))
        .get("total")
        or ZERO_DECIMAL
    )
    pending_reminders = ReminderTask.objects.filter(status=ReminderTask.Status.PENDING).order_by("remind_at", "id")
    course_reminders = list(pending_reminders.filter(reminder_type=ReminderTask.ReminderType.COURSE)[:6])
    receivable_reminders = list(pending_reminders.filter(reminder_type=ReminderTask.ReminderType.RECEIVABLE)[:6])
    recent_notes = list(SessionNote.objects.select_related("student", "schedule").order_by("-created_at")[:4])

    finance_chart = _build_finance_chart(today)
    student_overview = _build_student_overview()

    today_total = len(today_schedules)
    today_totals = {
        "all": today_total,
        "pending": sum(1 for schedule in today_schedules if schedule.status == Schedule.Status.PENDING),
        "completed": sum(1 for schedule in today_schedules if schedule.status == Schedule.Status.COMPLETED),
        "canceled": sum(1 for schedule in today_schedules if schedule.status == Schedule.Status.CANCELED),
        "rescheduled": sum(1 for schedule in today_schedules if schedule.status == Schedule.Status.RESCHEDULED),
        "delta": today_total - yesterday_schedule_count,
    }

    context = {
        "today": today,
        "today_label": f"{today.month}月{today.day}日 {WEEKDAY_LABELS[today.weekday()]}",
        "today_schedules": today_schedules[:5],
        "today_totals": today_totals,
        "active_students_count": active_students_count,
        "lead_students_count": lead_students_count,
        "outstanding_total": outstanding_total,
        "open_receivables_count": len(open_receivables),
        "month_received_total": month_received_total,
        "month_received_count": PaymentEntry.objects.filter(received_at__date__gte=month_start, received_at__date__lte=today).count(),
        "pending_reminders_count": pending_reminders.count(),
        "course_reminders": course_reminders,
        "receivable_reminders": receivable_reminders,
        "student_overview": student_overview,
        "recent_notes": recent_notes,
        "finance_chart": finance_chart,
        "receivables": open_receivables[:6],
    }
    return render(request, "dashboard/home.html", context)


@login_required
def report_center(request):
    return render(request, "dashboard/report_center.html")


@login_required
def export_students_csv(request):
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="students.csv"'
    writer = csv.writer(response)
    writer.writerow(["姓名", "昵称", "手机号", "状态", "服务类型", "标签"])
    for student in Student.objects.order_by("name"):
        writer.writerow([student.name, student.nickname, student.phone, student.get_status_display(), student.service_type, student.tags])
    return response


@login_required
def export_receivables_csv(request):
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="receivables.csv"'
    writer = csv.writer(response)
    writer.writerow(["学员", "标题", "应收金额", "实收汇总", "状态", "应收日期", "到期日期"])
    for receivable in Receivable.objects.select_related("student").order_by("due_date"):
        writer.writerow(
            [
                receivable.student.name,
                receivable.title,
                receivable.amount_due,
                receivable.amount_received,
                receivable.get_status_display(),
                receivable.issue_date,
                receivable.due_date,
            ]
        )
    return response


def _build_report_workbook(title: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "报表"
    sheet["A1"] = title
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["A2"] = f"生成时间：{timezone.localtime():%Y-%m-%d %H:%M}"
    fill = PatternFill(fill_type="solid", fgColor="DCEAFE")
    for cell in ("A1", "A2"):
        sheet[cell].fill = fill
    return workbook, sheet


@login_required
def export_hours_report(request):
    workbook, sheet = _build_report_workbook("学员课时报表模板")
    headers = ["学员", "服务方案", "排课开始", "状态", "计划时长", "实际时长", "扣减课时", "新增欠课时"]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=4, column=index, value=header)
    for row_index, schedule in enumerate(
        Schedule.objects.select_related("student", "service_plan").order_by("start_at"),
        start=5,
    ):
        completion = getattr(schedule, "completion", None)
        sheet.cell(row=row_index, column=1, value=schedule.student.name)
        sheet.cell(row=row_index, column=2, value=str(schedule.service_plan))
        sheet.cell(row=row_index, column=3, value=timezone.localtime(schedule.start_at).strftime("%Y-%m-%d %H:%M"))
        sheet.cell(row=row_index, column=4, value=schedule.get_status_display())
        sheet.cell(row=row_index, column=5, value=float(schedule.duration_hours))
        sheet.cell(row=row_index, column=6, value=float(completion.actual_duration_hours) if completion else "")
        sheet.cell(row=row_index, column=7, value=float(completion.deducted_hours) if completion else "")
        sheet.cell(row=row_index, column=8, value=float(completion.owed_hours_added) if completion else "")
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="hours-report-template.xlsx"'
    return response


@login_required
def export_finance_report(request):
    workbook, sheet = _build_report_workbook("经营收款报表模板")
    headers = ["学员", "标题", "应收金额", "实收汇总", "未收金额", "状态", "应收日期", "到期日期"]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=4, column=index, value=header)
    for row_index, receivable in enumerate(Receivable.objects.select_related("student").order_by("due_date"), start=5):
        sheet.cell(row=row_index, column=1, value=receivable.student.name)
        sheet.cell(row=row_index, column=2, value=receivable.title)
        sheet.cell(row=row_index, column=3, value=float(receivable.amount_due))
        sheet.cell(row=row_index, column=4, value=float(receivable.amount_received))
        sheet.cell(row=row_index, column=5, value=float(receivable.outstanding_amount))
        sheet.cell(row=row_index, column=6, value=receivable.get_status_display())
        sheet.cell(row=row_index, column=7, value=receivable.issue_date.strftime("%Y-%m-%d"))
        sheet.cell(row=row_index, column=8, value=receivable.due_date.strftime("%Y-%m-%d"))
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="finance-report-template.xlsx"'
    return response
